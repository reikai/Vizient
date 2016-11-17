#! python3

# backlog.py - Adjusts the backlog report and gets it ready to make a pivot table.

import openpyxl
from openpyxl.worksheet import get_column_letter
from openpyxl.worksheet import column_index_from_string
import os
from datetime import date

# It will be saved in the STBDdir, where the rest of the reports are.
print('Changing dir...')
STBDdir = '\\\\filecluster01\\dfs\\VhaSecure1\\ServiceDeliveryFinancials\\Performance Services\\_2016\\Monthly Financials'
os.chdir(STBDdir)

# Make a list of excel files, then choose the latest modified.
print('Getting a list of files, then sorting them in descending order...')
files = []
for f in os.listdir(os.getcwd()):                # Get all of the '.xlsx' files into the list 'files'.
    if f.endswith('.xlsx'):
        files.append(f)

files.sort(key=lambda x: os.path.getmtime(x))    # This lines sorts them ascending.
files = list(reversed(files))                    # This reverses the sort. After this line, list 'files' is descending.
latestfile = files[0]                            # The latest file is a string.

# Load up the STBD report as the workbook, and put the only sheet as the active one.
print('Loading wb and sheet...')
wb = openpyxl.load_workbook(latestfile)
sheet = wb.active

# Move all data from the 2nd row up, deleting the first row.
for row in range(2, sheet.max_row):
    for cell in range(1, sheet.max_column + 1):
        sheet[get_column_letter(cell) + str(row - 1)] = sheet[get_column_letter(cell) + str(row)].value

# Delete the last cell in 'A'. It just has a date of refresh.
while True:
    bottomA = sheet.max_row
    if sheet['A' + str(bottomA)] == None:
        bottomA -= 1
    else:
        sheet['A' + str(bottomA)] = None
        break

# This date is to save the file with a standard name.
d = date.today()
d = d.__str__()

wb.save(d + ' STBD Report.xlsx')



