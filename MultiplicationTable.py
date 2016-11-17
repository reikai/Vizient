#! python3

# Takes a number from a command line and makes a multiplication table from 1 to the given number.
# The Excel workbook is saved in the CWD.

import os
import sys
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font

# Get length of table.
lastNumber = int(sys.argv[1])

# Create a new workbook, and name the sheet according to the command line.
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Multiplication Table'

def makeTable(tableLength):
    # tableLength is an integer.

    # Make the outside rows, skipping the first cell.
    for c in range(2, tableLength + 2):
        sheet[get_column_letter(c) + '1'] = c - 1
        sheet['A' + str(c)] = c - 1

    # Use multiplication to fill in the table.
    for row in range(2, tableLength + 2):
        for column in range(2, tableLength + 2):
            sheet[get_column_letter(column) + str(row)] = sheet['A' + str(row)].value * sheet[get_column_letter(column) + '1'].value


def Formatting(rows, columns):
    sheet.freeze_panes = 'B2'

    segoeUIFontObj = Font(name='Segoe UI', size=10)

    for row in range(1, rows + 1):
        for column in range(1, columns + 1):
            sheet[get_column_letter(column) + str(row)].font = segoeUIFontObj


# Function Calling Zone
Formatting(lastNumber, lastNumber)
makeTable(lastNumber)


# Save workbook
wb.save('Multiplication Table.xlsx')

