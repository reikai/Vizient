#! python3

# Gives statistics on sample Census data

import openpyxl
from openpyxl.cell import get_column_letter

def main():
    print('Opening workbook...')
    wb = openpyxl.load_workbook('censuspopdata.xlsx')
    sheet = wb.get_sheet_by_name('Population by Census Tract')

    # Get the Generator Excel object (range of cells with data)
    LastRow = sheet.max_row
    LastColumn = get_column_letter(sheet.max_column)

    generator = sheet['A1': LastColumn + str(LastRow)]

    # Get a unique list of counties
    CountyList = []
    for row in generator:
        if row[2].value not in CountyList:
            if row[2].value == 'County':  # We don't want to include the header row.
                continue
            CountyList = list(set(CountyList))
            CountyList.append(row[2].value)

    # Make a dictionary item for each county with lists for the county values
    # The list will hold population, tract number, and state

    CountyDict = {}  # This dictionary is unique, due to the above exercise
    for county in CountyList:
        CountyDict[county] = [0, [], '']  # County = [pop, [tract list], 'STATE']

    # Generate Excel Data Range and sum up population, create tract list, and add state for each county
    generator = sheet['A1': LastColumn + str(LastRow)]
    i = 0                                                 # Counter for skipping header row.
    for row in generator:
        if i == 0:
            i += 1
            continue
        CountyDict[row[2].value][0] += row[3].value       # Add the population to the right county.
        CountyDict[row[2].value][1].append(row[0].value)  # Append a tract number to the Tract list for that county.
        CountyDict[row[2].value][2] = row[1].value        # Add the State 2-letter abbreviation.


    # Print list of counties, number of tracts in each county, population for each county.

    for key in CountyDict.keys():
        print('County ' + key + ' has ' + str(len(CountyDict[key][1])) + ' tracts and ' + str(CountyDict[key][0]) + ' population.')

main()
